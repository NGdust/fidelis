"""Tests for the fidelis.sources input adapters.

Cover from_records, from_csv, from_json, from_excel and the SourceData contract:
lazy records, field inventory, sample, raw_kind and metadata.

Each test is independent — it does not rely on state left behind by another.
"""

from __future__ import annotations

import io
import json

import pytest

import fidelis as fp
from fidelis.sources import DEFAULT_SAMPLE_SIZE, SourceData
from fidelis.sources.base import (
    collect_field_names,
    materialize_sample,
)

from conftest import User, Product, USER_MAPPINGS  # noqa: F401  (import availability)


# --------------------------------------------------------------------------- #
# SourceData / base helpers
# --------------------------------------------------------------------------- #

def test_sourcedata_iter_delegates_to_records():
    sd = SourceData(records=[{"a": 1}, {"a": 2}], field_names=["a"], sample=[{"a": 1}])
    assert list(sd) == [{"a": 1}, {"a": 2}]
    # raw_kind defaults to structured, meta is an empty dict
    assert sd.raw_kind == "structured"
    assert sd.meta == {}


def test_collect_field_names_union_preserves_first_seen_order():
    rows = [
        {"a": 1, "b": 2},
        {"b": 3, "c": 4},
        {"a": 5, "d": 6},
    ]
    assert collect_field_names(rows) == ["a", "b", "c", "d"]


def test_materialize_sample_reiterable_tail():
    rows = ({"i": i} for i in range(5))  # a real generator
    sample, records = materialize_sample(rows, sample_size=2)
    assert sample == [{"i": 0}, {"i": 1}]
    # the tail together with the sample is fully iterable again
    assert list(records) == [{"i": i} for i in range(5)]


def test_materialize_sample_smaller_than_size():
    sample, records = materialize_sample([{"x": 1}], sample_size=10)
    assert sample == [{"x": 1}]
    assert list(records) == [{"x": 1}]


# --------------------------------------------------------------------------- #
# from_records
# --------------------------------------------------------------------------- #

def test_from_records_basic_structured(user_records):
    sd = fp.from_records(user_records)
    assert isinstance(sd, SourceData)
    assert sd.raw_kind == "structured"
    assert sd.field_names == ["E-mail", "Name", "Date"]
    assert sd.sample == user_records
    assert sd.meta == {"adapter": "records", "row_count": 2}


def test_from_records_field_names_union_over_heterogeneous_dicts():
    rows = [
        {"a": 1},
        {"b": 2, "c": 3},
        {"a": 4, "d": 5},
    ]
    sd = fp.from_records(rows)
    # union of keys across all records in first-seen order
    assert sd.field_names == ["a", "b", "c", "d"]
    assert list(sd.records) == rows


def test_from_records_accepts_generator():
    sd = fp.from_records({"n": i} for i in range(3))
    assert sd.field_names == ["n"]
    assert sd.meta["row_count"] == 3
    assert list(sd.records) == [{"n": 0}, {"n": 1}, {"n": 2}]


def test_from_records_lazy_sample_default_size():
    rows = [{"id": i} for i in range(DEFAULT_SAMPLE_SIZE + 7)]
    sd = fp.from_records(rows)
    # the sample is capped at DEFAULT_SAMPLE_SIZE...
    assert len(sd.sample) == DEFAULT_SAMPLE_SIZE
    assert sd.sample == rows[:DEFAULT_SAMPLE_SIZE]
    # ...but a full pass loses no record (sample + tail)
    assert list(sd.records) == rows


def test_from_records_custom_sample_size():
    rows = [{"id": i} for i in range(10)]
    sd = fp.from_records(rows, sample_size=3)
    assert len(sd.sample) == 3
    assert list(sd.records) == rows


def test_from_records_empty():
    sd = fp.from_records([])
    assert sd.field_names == []
    assert sd.sample == []
    assert list(sd.records) == []
    assert sd.meta["row_count"] == 0


# --------------------------------------------------------------------------- #
# from_csv — different inputs
# --------------------------------------------------------------------------- #

def test_from_csv_inline_text(user_csv_text):
    sd = fp.from_csv(user_csv_text)
    assert sd.raw_kind == "text"
    assert sd.field_names == ["E-mail", "Name", "Date"]
    assert sd.meta["adapter"] == "csv"
    # inline text is not a file, so there is no source name
    assert sd.meta["source_name"] is None
    assert sd.meta["row_count"] == 2
    recs = list(sd.records)
    assert recs[0] == {"E-mail": " A@B.COM ", "Name": " Alice ", "Date": "01.02.2026"}


def test_from_csv_file_path_str(tmp_path):
    p = tmp_path / "data.csv"
    p.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
    sd = fp.from_csv(str(p))
    assert sd.raw_kind == "text"
    assert sd.field_names == ["a", "b"]
    assert sd.meta["source_name"] == str(p)
    assert list(sd.records) == [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]


def test_from_csv_pathlike(tmp_path):
    p = tmp_path / "data.csv"
    p.write_text("x,y\n10,20\n", encoding="utf-8")
    sd = fp.from_csv(p)  # Path is an os.PathLike
    assert sd.field_names == ["x", "y"]
    assert sd.meta["source_name"] == str(p)
    assert list(sd.records) == [{"x": "10", "y": "20"}]


def test_from_csv_raw_bytes():
    sd = fp.from_csv(b"a,b\n1,2\n")
    assert sd.raw_kind == "text"
    assert sd.field_names == ["a", "b"]
    assert sd.meta["source_name"] is None
    assert list(sd.records) == [{"a": "1", "b": "2"}]


def test_from_csv_bytes_buffer():
    buf = io.BytesIO(b"a,b\n1,2\n3,4\n")
    sd = fp.from_csv(buf)
    assert sd.field_names == ["a", "b"]
    assert list(sd.records) == [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]


def test_from_csv_text_buffer():
    buf = io.StringIO("a,b\n1,2\n")
    sd = fp.from_csv(buf)
    assert sd.field_names == ["a", "b"]
    assert list(sd.records) == [{"a": "1", "b": "2"}]


def test_from_csv_buffer_name_propagated(tmp_path):
    p = tmp_path / "named.csv"
    p.write_text("a,b\n1,2\n", encoding="utf-8")
    with open(p, "rb") as fh:
        sd = fp.from_csv(fh)
    # the file object's name ends up in meta.source_name
    assert sd.meta["source_name"] == str(p)


# --------------------------------------------------------------------------- #
# from_csv — dialect autodetection
# --------------------------------------------------------------------------- #

@pytest.mark.parametrize(
    "delim",
    [",", ";", "\t"],
)
def test_from_csv_delimiter_autodetection(delim):
    text = delim.join(["col1", "col2", "col3"]) + "\n"
    text += delim.join(["1", "2", "3"]) + "\n"
    text += delim.join(["4", "5", "6"]) + "\n"
    sd = fp.from_csv(text)
    assert sd.meta["delimiter"] == delim
    assert sd.field_names == ["col1", "col2", "col3"]
    assert list(sd.records) == [
        {"col1": "1", "col2": "2", "col3": "3"},
        {"col1": "4", "col2": "5", "col3": "6"},
    ]


def test_from_csv_explicit_delimiter_overrides_autodetect():
    # semicolon in the data, but we request an explicit delimiter over autodetect
    sd = fp.from_csv("a;b\n1;2\n", delimiter=";")
    assert sd.meta["delimiter"] == ";"
    assert list(sd.records) == [{"a": "1", "b": "2"}]


# --------------------------------------------------------------------------- #
# from_csv — encoding detection
# --------------------------------------------------------------------------- #

def test_from_csv_utf8_encoding_detected():
    # Cyrillic values are intentional test data to exercise non-ASCII decoding.
    text = "name;city\nЖеня;Уфа\n"
    sd = fp.from_csv(text.encode("utf-8"))
    assert sd.meta["encoding"] in ("utf-8", "utf-8-sig")
    assert list(sd.records) == [{"name": "Женя", "city": "Уфа"}]


def test_from_csv_cp1251_encoding_detected():
    # Cyrillic values are intentional test data for the cp1251 codepage decode path.
    text = "name;city\nИван;Москва\nПётр;Казань\n"
    raw = text.encode("cp1251")
    # sanity: these bytes do not decode as utf-8
    with pytest.raises(UnicodeDecodeError):
        raw.decode("utf-8")
    sd = fp.from_csv(raw)
    assert sd.meta["encoding"] == "cp1251"
    assert list(sd.records) == [
        {"name": "Иван", "city": "Москва"},
        {"name": "Пётр", "city": "Казань"},
    ]


def test_from_csv_explicit_encoding_overrides_detect():
    # Cyrillic values are intentional test data for the cp1251 codepage decode path.
    raw = "name;city\nИван;Москва\n".encode("cp1251")
    sd = fp.from_csv(raw, encoding="cp1251")
    assert sd.meta["encoding"] == "cp1251"
    assert list(sd.records)[0] == {"name": "Иван", "city": "Москва"}


# --------------------------------------------------------------------------- #
# from_csv — ragged rows
# --------------------------------------------------------------------------- #

def test_from_csv_ragged_short_row_padded_with_none():
    sd = fp.from_csv("a,b,c\n1,2\n")
    # a short row is padded with None, no data is lost
    assert list(sd.records) == [{"a": "1", "b": "2", "c": None}]


def test_from_csv_ragged_long_row_truncated():
    sd = fp.from_csv("a,b,c\n1,2,3,4,5\n")
    # extra cells are truncated to the header width
    assert list(sd.records) == [{"a": "1", "b": "2", "c": "3"}]


def test_from_csv_blank_rows_skipped():
    sd = fp.from_csv("a,b\n1,2\n\n3,4\n")
    assert list(sd.records) == [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]
    assert sd.meta["row_count"] == 2


def test_from_csv_header_whitespace_stripped():
    sd = fp.from_csv(" a , b \n1,2\n")
    assert sd.field_names == ["a", "b"]


# --------------------------------------------------------------------------- #
# from_json
# --------------------------------------------------------------------------- #

def test_from_json_from_list():
    data = [{"a": 1, "b": 2}, {"a": 3, "b": 4}]
    sd = fp.from_json(data)
    assert sd.raw_kind == "structured"
    assert sd.field_names == ["a", "b"]
    assert list(sd.records) == data
    assert sd.meta == {"adapter": "json", "row_count": 2}


def test_from_json_from_json_string():
    sd = fp.from_json('[{"x": 1}, {"x": 2}]')
    assert sd.field_names == ["x"]
    assert list(sd.records) == [{"x": 1}, {"x": 2}]


def test_from_json_from_bytes():
    sd = fp.from_json(b'[{"x": 1}]')
    assert list(sd.records) == [{"x": 1}]


def test_from_json_from_file_path(tmp_path):
    p = tmp_path / "data.json"
    p.write_text(json.dumps([{"a": 1}, {"a": 2}]), encoding="utf-8")
    sd = fp.from_json(str(p))
    assert sd.field_names == ["a"]
    assert list(sd.records) == [{"a": 1}, {"a": 2}]


def test_from_json_from_pathlike(tmp_path):
    p = tmp_path / "data.json"
    p.write_text(json.dumps([{"k": "v"}]), encoding="utf-8")
    sd = fp.from_json(p)
    assert list(sd.records) == [{"k": "v"}]


def test_from_json_single_object_wrapped():
    sd = fp.from_json({"only": 1})
    assert sd.field_names == ["only"]
    assert list(sd.records) == [{"only": 1}]
    assert sd.meta["row_count"] == 1


def test_from_json_single_object_from_string():
    sd = fp.from_json('{"only": 1}')
    assert list(sd.records) == [{"only": 1}]


def test_from_json_field_names_union_heterogeneous():
    sd = fp.from_json([{"a": 1}, {"b": 2}, {"a": 3, "c": 4}])
    assert sd.field_names == ["a", "b", "c"]


def test_from_json_non_object_array_element_raises():
    with pytest.raises(ValueError):
        fp.from_json([{"a": 1}, 2, {"a": 3}])


def test_from_json_array_of_scalars_raises():
    with pytest.raises(ValueError):
        fp.from_json([1, 2, 3])


def test_from_json_unsupported_input_type_raises():
    with pytest.raises(TypeError):
        fp.from_json(12345)


def test_from_json_lazy_sample_default_size():
    data = [{"i": i} for i in range(DEFAULT_SAMPLE_SIZE + 5)]
    sd = fp.from_json(data)
    assert len(sd.sample) == DEFAULT_SAMPLE_SIZE
    assert list(sd.records) == data


# --------------------------------------------------------------------------- #
# from_excel
# --------------------------------------------------------------------------- #

def _make_xlsx(path, sheets):
    """Create a real .xlsx. sheets: list[(title, rows)]."""

    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    for idx, (title, rows) in enumerate(sheets):
        ws = default if idx == 0 else wb.create_sheet()
        ws.title = title
        for row in rows:
            ws.append(row)
    wb.save(path)


def test_from_excel_header_and_rows(tmp_path):
    p = tmp_path / "book.xlsx"
    _make_xlsx(
        p,
        [("Sheet1", [["sku", "price", "in_stock"], ["A1", 9.99, True], ["B2", 5.0, False]])],
    )
    sd = fp.from_excel(p)
    assert sd.raw_kind == "structured"
    assert sd.field_names == ["sku", "price", "in_stock"]
    assert sd.meta["adapter"] == "excel"
    assert sd.meta["row_count"] == 2
    recs = list(sd.records)
    assert recs == [
        {"sku": "A1", "price": 9.99, "in_stock": True},
        {"sku": "B2", "price": 5.0, "in_stock": False},
    ]


def test_from_excel_sheet_selection_by_name(tmp_path):
    p = tmp_path / "multi.xlsx"
    _make_xlsx(
        p,
        [
            ("First", [["a"], [1], [2]]),
            ("Second", [["b"], [9], [8]]),
        ],
    )
    sd = fp.from_excel(p, sheet="Second")
    assert sd.field_names == ["b"]
    assert list(sd.records) == [{"b": 9}, {"b": 8}]
    assert sd.meta["sheet"] == "Second"


def test_from_excel_sheet_selection_by_index(tmp_path):
    p = tmp_path / "multi.xlsx"
    _make_xlsx(
        p,
        [
            ("First", [["a"], [1]]),
            ("Second", [["b"], [9]]),
        ],
    )
    sd = fp.from_excel(p, sheet=1)
    assert sd.field_names == ["b"]
    assert list(sd.records) == [{"b": 9}]
    assert sd.meta["sheet"] == 1


def test_from_excel_default_active_sheet(tmp_path):
    p = tmp_path / "active.xlsx"
    _make_xlsx(p, [("Only", [["c1", "c2"], ["x", "y"]])])
    sd = fp.from_excel(p)
    assert sd.field_names == ["c1", "c2"]
    assert list(sd.records) == [{"c1": "x", "c2": "y"}]
    assert sd.meta["sheet"] is None


def test_from_excel_blank_rows_skipped(tmp_path):
    p = tmp_path / "blanks.xlsx"
    _make_xlsx(
        p,
        [("S", [["a", "b"], [1, 2], [None, None], [3, 4]])],
    )
    sd = fp.from_excel(p)
    assert list(sd.records) == [
        {"a": 1, "b": 2},
        {"a": 3, "b": 4},
    ]
    assert sd.meta["row_count"] == 2


def test_from_excel_short_row_padded_with_none(tmp_path):
    p = tmp_path / "short.xlsx"
    _make_xlsx(p, [("S", [["a", "b", "c"], [1, 2]])])
    sd = fp.from_excel(p)
    assert list(sd.records) == [{"a": 1, "b": 2, "c": None}]


def test_from_excel_missing_header_cell_gets_col_name(tmp_path):
    from openpyxl import Workbook

    p = tmp_path / "gap.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["a", None, "c"])
    ws.append([1, 2, 3])
    wb.save(p)
    sd = fp.from_excel(p)
    assert sd.field_names == ["a", "col_1", "c"]
    assert list(sd.records) == [{"a": 1, "col_1": 2, "c": 3}]


def test_from_excel_empty_sheet(tmp_path):
    p = tmp_path / "empty.xlsx"
    _make_xlsx(p, [("S", [])])
    sd = fp.from_excel(p)
    assert sd.field_names == []
    assert list(sd.records) == []
    assert sd.raw_kind == "structured"


def test_from_excel_lazy_sample_default_size(tmp_path):
    p = tmp_path / "big.xlsx"
    rows = [["id"]] + [[i] for i in range(DEFAULT_SAMPLE_SIZE + 4)]
    _make_xlsx(p, [("S", rows)])
    sd = fp.from_excel(p)
    assert len(sd.sample) == DEFAULT_SAMPLE_SIZE
    assert len(list(sd.records)) == DEFAULT_SAMPLE_SIZE + 4


# --------------------------------------------------------------------------- #
# Cross-adapter raw_kind contract
# --------------------------------------------------------------------------- #

def test_raw_kind_text_only_for_csv(tmp_path):
    csv_sd = fp.from_csv("a,b\n1,2\n")
    rec_sd = fp.from_records([{"a": 1}])
    json_sd = fp.from_json([{"a": 1}])

    xp = tmp_path / "x.xlsx"
    _make_xlsx(xp, [("S", [["a"], [1]])])
    xl_sd = fp.from_excel(xp)

    assert csv_sd.raw_kind == "text"
    assert rec_sd.raw_kind == "structured"
    assert json_sd.raw_kind == "structured"
    assert xl_sd.raw_kind == "structured"


@pytest.mark.parametrize(
    "factory",
    [
        lambda tp: fp.from_csv("a,b\n1,2\n"),
        lambda tp: fp.from_records([{"a": 1, "b": 2}]),
        lambda tp: fp.from_json([{"a": 1, "b": 2}]),
    ],
)
def test_all_adapters_populate_field_names_and_sample(factory, tmp_path):
    sd = factory(tmp_path)
    assert sd.field_names  # non-empty field inventory
    assert sd.sample  # non-empty sample
    assert isinstance(sd, SourceData)
