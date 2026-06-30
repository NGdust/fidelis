"""Excel adapter (xlsx/xls) — a structured input, no Stage 1.

Requires the optional ``openpyxl`` dependency (extra ``fidelis[excel]``).
"""

from __future__ import annotations

import os
from typing import Optional, Union

from .base import DEFAULT_SAMPLE_SIZE, SourceData, materialize_sample

PathOrBuffer = Union[str, os.PathLike, "os.PathLike"]


def from_excel(
    path_or_buffer: PathOrBuffer,
    *,
    sheet: Optional[Union[str, int]] = None,
    sample_size: int = DEFAULT_SAMPLE_SIZE,
) -> SourceData:
    """Accept xlsx/xls. The first row of the sheet is treated as the header.

    Args:
        sheet: Sheet name or index. Defaults to the active sheet.
    """

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on the environment
        raise ImportError(
            "from_excel requires openpyxl. Install: pip install 'fidelis[excel]'"
        ) from exc

    wb = load_workbook(path_or_buffer, read_only=True, data_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return SourceData(records=[], field_names=[], sample=[], raw_kind="structured")

        header = [
            (str(h).strip() if h is not None else f"col_{i}")
            for i, h in enumerate(header_row)
        ]
        records: list[dict] = []
        for cells in rows_iter:
            if cells is None or not any(c is not None for c in cells):
                continue
            padded = list(cells) + [None] * (len(header) - len(cells))
            records.append(dict(zip(header, padded[: len(header)])))
    finally:
        wb.close()

    sample, materialized = materialize_sample(records, sample_size)
    return SourceData(
        records=materialized,
        field_names=header,
        sample=sample,
        raw_kind="structured",
        meta={"adapter": "excel", "sheet": sheet, "row_count": len(records)},
    )
